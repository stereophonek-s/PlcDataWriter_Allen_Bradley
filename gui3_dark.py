import tkinter as tk
from tkinter import filedialog, messagebox
from pycomm3 import LogixDriver
from openpyxl import load_workbook
import traceback
import base64
from io import BytesIO
from PIL import Image, ImageTk

# Define your static PLC address
PLC_ADDRESS = '192.168.33.100'

# Define the static starting points for each column
START_CELL_LENGTH = 'J2'
START_CELL_QTY = 'I2'
START_CELL_STRING = 'B2'

your_icon_base64_code = 'AAABAAEAAAAAAAEAIADkEQAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAABGeSURBVHja7Z15kBxVHcenu+eezR7ZzWaT7IYc5IRcZJMYD0qBIMSjpAIlKpDEBDnEUjECcpYKBrFAKEpL+YfyKqTUUjnkP/Aqj7JQUMGoSLS8wKNUUDAJ+Pz9et6SZKZ7rp3Z7Z7+fKo+FarYnZ3ufr9vd79+73UqBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAE0wJG4UXy+eLb4FMcGebWthgzjYzYV/rHit+EPxb+Lz4gvii4gJ9gVbC1oT3xevE5d1U+E74hvFR0WDiHV9XDzL1k7s0Q15moOK2JRPiWfEvfjXir/iYCK2fCVwfFyLPyN+MnDDHNc42X7jFoYRE6/WgtZESAjcLqbjGADLxd9WbpCT7TPZsa2msOpSU1x7JWLi1VrIjp1unExvUADsj2un4FvFQ0dtjJs1uUVnmdKGvaa0cW/5X8Ska2sht3Cb1EimMgAO2keFseOqyjRzS6OmuO5qu+EfQcSX3Cu1cZVxi3ODrgIuj2MA3Fy5IV7/clNa/yEONmKQ6z9ovL6lQQHw0TgGwK2BATBOACCGB8CyoAD4WBwD4ONcASASAAQAIgHQjgCgxxjjIAHQ3gDYeKMpjV/vP0EorrnMFFfvQYye2jb1Kdf4DeU2SwBMNgD2+jslv3SHSc/aaNzSPOPmh4yTG0CMnNo29VG3ttX8ku1+223qioAAqHguKqmaHlpnHC/HuHCMldpm00PrpQ2/v/EQIAAOF78Oj/RmLKIxYaz1ehc3HgIEgD3zn3CdSc9cTQPCrjA9uM5v03VDgAAoB0Bu8dlBY6KP0EGMoCHt1U0fngNDAHyo/k4YWBm8I72scbIl4+R6EKOntE1to0FhoIVd7hQkAOre+/vzo6uKP2ecQr84gBhhpe2mqzut9SlBYdX7al8FJD4ANt5o8svPr+71dxzZgTNoXBgPpa1qmz36qUDe5FdcUHt8AAEgAbDkPP+eqWoFoVwvDQtjEgC9VQGgfVo6noUAaCQAHAIA4x4AbkAAbCcAuAJAAoAAIACQACAACAAkAAgAAgAJAAKAAEACgAAgAJAAIAAIACQACAACoCPDVZ282mftZ5g1AUAAdK1a5Nke42QK5ckqun9dr8J0ebKV/Iz/s/o77DsCgACIa9H3H56d5jdGp4k57E75dyZmX+a5OiAACID4FH6maPeh04YFLRz/s/QzCQICgACIsnLpXrXv2qkGgd4esK8JAAIgYmf9dL56Blq9hSutTYWATs2Wv8XVAAFAAESkg6+8Ck3twi2Ky8Vt4uXiLeKnrDeLl9n/pz9TaCQItH+AjkICgACY5oZWc+3ElFkoXiTeI/5OPCCaEA/Yn/m6eLG4uJFbAo4LAUAATMeZv7fm/f5s8VLx5+KLNYo+TP2dx+1nDNcLAa4ECAACYGrv+VNeJvS+fov4TfGFFgq/Uv2Mh+xnhvYXyHehT4AAIACmSu3wCyjErHih+Mc2FH6lf7a3EpmwENCOQY4NAUAATMGjvoDefi3MPeKzHSj+CZ+xfyMb9nSAR4QEAAHQ4Uv/gPt+x575n+lg8U/4rP1bgbcDLrcCBAAB0Dl1hF9A4b1a/MMUFP+Ef7J9AoF9EDpikAAgAAiADjzvDzj7z7GddGaK/Zb928FPBfoJAAKAAGjvvX8pcGz/FeL/piEA9DHhB0LmDvgTiAgAAoAAaJ9Bo/10kM9j01D8E+4Tjw0bJUgAEAAEQAtj+vUVaHoG1bn7+mhN3zPnT+mtPvtf3OIgn3ZeBVwU8kTADwH97roN/hoDpfIxTMLtAQFAADS3WEep/EJJXZijwQk9JfH+aSz+Ce+x8wwanUjkH1vZ1vI6A30EAAGQ0ADQxq9neS36Fubtr+rQgJ9m/b24suV1Brzy1UG3BQEBQADUXbDD8SY1N/9c8VAEAuCg+JbJrjPgeN214AgBQAAEN4wZoWP4m7FH/HIEin/Cu+0tyaQXHNE5Bd3wGngCgAAIfIxX2SgC1CG9I+Jq8WTxDDtP/1Rxs/gq8VbxuQgFgH6X28RX2u94qv3ObxJPsrcrI7XmEVQe67g/RiQACICqEXw1Ovc8uxDHO8TPiY+KT9lhtwfsZfa/xX+I/4pQ4Vf6T/sd/22/8wG7Dbotj4ifF88Xl9ltrrnqkPYNxHWpcgKAADhc/IXQTj4dT7/GntH3t2n6btTVbXzSbvO6mkuSTYQAAUAAxDUA/LH7wcU/YEfxPZmAog/zt3Z5soFaVwJxvB0gAAgAv8Mv5J5fl9q6KyFn/Hoesp2IS2r1CcStY5AASHgA+BN3gnv79bL3OxEoPL3y+Iz4WXsmnu7v821xbbcsO0YAJDwA/Pv+6oa8QvxeBIpNZw+Oi67tjNsUke/1A/G40BWHCgQAARDTgy/OssNmp7vIHrS98JXfb3VEQuBr4mDorUAvAUAARHzWno7pr9huPdNeN80Td2oVf5RCQPtFrrH7rPoqIEcAEAARX6o74Oz/8giM269X/FEKAV2A9MQ4XwUQAAkNgIAlu3QBzTtjUvxRCoEvhryhKBZjAwiAJAZAf2DP/yY7Ei6Kxe/lXONl3UiGwN/DrgJ04dGojxAkABIYACGdf9dHtPizPWmz8YrlZv2lS0264EUyBD4SdhuQ7yUACIAortl39PbOnMYCqlX8GSn+l129wuzcd5rZ8fhrzfo9yyIZAjpeoj/oNiDqowMJgAQGQMCz//XiXyJc/G9/4nTf7Y9FMwSeChscFPV+AAIgeQEQ9PjvTDsrLrLF/2trREPgv3Y6dCpui44SAAkMgICFPt4Th+KPeAhcErZwSJQ7AgmApAVA8BOAD8el+CMcAtfG8UkAAZDEAKh+a89NcSr+iIbAjQQAARDXK4Cr41b8EQyBD3ILQADEtQ/ggg6/tqsjxR+xELiETkACIK5PAd5ge7JjV/wRCQFdbPR1cZwURAAkcRxA9TyAlfbFGbEs/giEgC5UsjRwHECRACAAIrgEWMX6f7pW/jemsPhzfRmz+bqVfuHu3r/V7HqybKu3ABO/r5+185enmQ2XL/cDZqpC4IHA147pOoE9BAABEMVlwKrPkGeJ99rFQCrVcPhZE2/4qVX8jpMy808aNid/4gSz5Y71Zsun1/v/vuFLm825j2xpuhNQf0d/98jPOun2dWbuK4baMotQ5/3/3L7fMGjf3CeeE/YWoagvD0YAJHM6cNhiIDkxX2HOTnedZ1+qUW+xkB/XewefY2f32Rl+vvLfelWwZNuoeduPTmn4MeA5D59ilr15zP/dys8Lmz044Vr7XoN6C4HeIo4dsS/yAXpxffU4AZDQ9QD8CUHNv+hzgfjYEa/crvRg2Gu4G9TxHHPiTavLtwN1AmDXb7aaV9+yxrgZt+W/d6H9zkHbotv4U3G01W2JwzLhBEBSVwQKHhBUz7SExtZ03pwvjXtXtljl9kzRzJvky0RXX7CooSuAXfIzq3YvnNTfmiO3Qvqddwdsi27j6bKt6VY+Oy6rAxMACQ2AkKcB9Ru2FLf2IehtRJhSNJMpyuN3LWw4AI7bsWByL/nU71prW0L6S+oa9d5/AoAA8K8CtAD0jK2NoKZeeVhrI73aWjTax1D3M93A25Djd082AJzG/raXa+wsnespD55qZD9pWMg+jc3rwwmABAfAEQXbkM0OadXf0X1Rw6DbkEkHgF5+1/m7zW9PfxP7iTcDEQBxf2/8VD2NCJiXMOkA8Mffs28JAAIglouTEgAEAAFAABAABAABQAAQAAQAAUAAEAAEAAFAABAABAABQAAQAAQAAUAAEAAEAAFAABAABAABQAAQAAQAAUAAEAAEAAFAALB/CQACIJYBoHP8d9mXgdZSFwQhAAgAAqDLAmDl9mPMeY+eas6t43k/PdWseNt8AoAAIAC6KQBKI3kze3ygIYuz8wQAAUAAdFMATFoCgAAgAGISAB4BQAAQAMldECRgefJJq0tysW8JAAIgJg3Q9dpX/BPLgbFvCQACIC5rEvaWVyjOFOTf1iz/bjF+a/IRAAQAIgFAACASAAQAIgFAACASAAQAIgFAACASAAQAIgFAACASAAQAIgFAACASAAQAYmvDsAmAtgaAU56NNolx7bXHu0ffVMw+N8lqW02lHAKgbQGAGHcJgCYCwCEAkABIZgAsf4dxvDwNBrtKx8v5bZsAqBUAG/aa4rqrZScspdFgV+n1Lpa2fZXfxgmA0AAoh0Bh1XtNZniz8XqOMW5pFDG2ahtOD28yhePfXbv4CYCjQ6A0foMpnnCtpOY1iPFV2rC25brFTwCEBAFi7G2wvRMAiAmWAEAkAAgARAKAAEAkAOIbALdUB8AKUxq/vjwYAg/bTEdR17iX417p+IfDAuCmKBT0bPHl4kkN+Brxy5Ub4vbMN/mlO01+2W60FlZcVH+ASBcWv26zbjtt4AiX7jBuaSwoAO62NdVI7W0WR9pd/K8Uvyv+U3xWfKYBD1ZtiE7rdbP+sEgUXf23YLwZi0zhuHclJAT2yrZeYrzeRf62l/cBlttDtnoacdkDDdac1uY/xO+LJ7ar+GeI9zOcs7OmB9fKJWAC+kfkMjc9uI5j3nnvE0vtCIAF4hPs0M7qFkcScCsgl/5rr/S3lWPecX8ljrUjABaL+9mhHQ6AwrBfHF0fAGuuMG5+mGPeeZ+wJ+8OBYDryX1Luq4pz/Nt5GeTYsp1CYCq9uTSNlqsm5BXwXcwABzHuP2DJj0yZrzZo9iEus/c3gECoHL7e2fSnlptT/1D5WXzpjIAvMHZJj222KRHF2Ezyj5zB4YIgMrtH5hFe2qxPXlDI9MUAKMcsFYkAEICgPbUkgRAtwSAPgV4aWRgFyrbVlz7AX9bCQACgAA4sgDyg6aw8iJTXL2nq82vuNC4uUECgAAgAI7ep55xsn3GyQ10t7KNuq0EQLvsfB/AMXZQwREfLgEwc5gD1s4ASPo4CAKg9QCQk3HVi0hSqX3ivHYEgE4seLT6gA1xwFoOgFkUfeVM0ZkEQJtPKD8RZ7VrLsA3qwJgRj87v9V7tuG5YYM3EvrSDNd4s+bSNloNgIBxJeJDtnYnjSd+pupFB/miSc9byAFoRdlvbt/M8mgvvXdLsrIPtAHTllpvS06hFBQAd4npds0IvKzqD3hp442McQBadmF5NJfcvyVaHc1G8bd+NTky3zjpTFAAXNPO9QBOsXONAzpuOAjt6MhJrhz/SQXAzOGgJwBaq1vaGQDaEfhw9W1AwaTnLuBAIE6HUnt6Kx5w9n+4E6sC7a1+dj3xOJCDgTjlZ399mlR99jd2PU6n3QEwLv6x6iogk6MvAHGqi3/2qNReNqj4nxZf0YlFQbVH8bagxzhOsYdbAcSpvPSXmgt5rHqHmO3UysDHVY8KnFjxt8+k5xECiJ0ufq21kOL/tbi608uDv1N8Pqg/wC3NMOk58zlIiJ1wzjFSY71hxa81efFUvB+gZC8zAr+IkysYb9Yc/zk3Bw2xPYN9tKa0tmqMqLxT7Jmql4TMFe8JH9rp+UnlD+/ktgCxxcJf4NeQf9avPXT8AXH+VL8paJFde9zUCgJ9TqnDXnXKoj/yTW4REDFErRGpFa0Z/xl//TkjD4pLp+t1YTrd8AvioboTPuy4bx226GQy/iMMRJwwUx7SOzE/pP4kqhfEr4jHTvc7AwfEa8W/MLMNcUp82tbczKi8CdizLy/8mvgfDhBiR9TautfWmpeKIL3iNttB+FcOGGJb1Kvrr4tnin2pGKCPCl8mXmnDYJ8NhOdS5bcGv4iIVR60z/P/Jv5C/Kp4hbgp1aYXfU4H+VT5JYU6l2CreLa4U9yFiC+509bG62ytjNraAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAgGD+D6zPd5p5wAQGAAAAAElFTkSuQmCC'


def set_icon(base64_icon):
    icon_data = base64.b64decode(base64_icon)
    image = Image.open(BytesIO(icon_data))
    photo = ImageTk.PhotoImage(image)
    root.tk.call('wm', 'iconphoto', root._w, photo)


def browse_file():
    file_path = filedialog.askopenfilename()
    file_path_entry.delete(0, tk.END)
    file_path_entry.insert(0, file_path)


def extract_string_data(file_path, end_string):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    string_data = []

    for i in range(2, end_string + 2):
        cell_value = sheet[f'B{i}'].value
        if cell_value is None or i > end_string:
            string_data.append(None)
        else:
            string_data.append(cell_value)

    # Fill remaining cells with empty strings
    while len(string_data) < 121:
        string_data.append("")

    return string_data


def check_and_write_to_plc():
    try:
        if automated_mode.get():
            with LogixDriver(PLC_ADDRESS) as plc:
                check1_value = plc.read('Profile_AUTO').value
                check2_value = plc.read('REM10:O.Data[0].2').value

            if check1_value == 1 or check2_value == 1:
                messagebox.showinfo("Write Results", "Auto mode is ON. No write is performed.")
            else:
                write_to_plc()
        else:
            write_to_plc()
            messagebox.showinfo("Success", "Data has been written to PLC successfully")
    except Exception as e:
        error_message = f"An error occurred: {str(e)}"
        messagebox.showerror("Error", error_message)
        traceback.print_exc()


def write_to_plc():
    file_path = file_path_entry.get()
    end_length = int(length_end_entry.get())
    end_qty = int(qty_end_entry.get())
    end_string = int(string_end_entry.get())

    l_range = [f'{START_CELL_LENGTH}:{chr(ord(START_CELL_LENGTH[0]))}{end_length}']
    m_range = [f'{START_CELL_QTY}:{chr(ord(START_CELL_QTY[0]))}{end_qty}']

    # Extract data from the Excel file
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    l_data = [cell[0].value for cell in sheet[l_range[0]]]
    m_data = [cell[0].value for cell in sheet[m_range[0]]]

    # Initialize lists for data
    massive1_data = [0] * 120
    massive2_data = [0] * 120
    string2_data = [None] * 121  # Initialize with None

    # Extract string data from the Excel file
    string_data = extract_string_data(file_path, end_string)

    # Fill the lists with values from Excel up to the specified endpoints
    for i, value in enumerate(l_data[:end_length]):
        massive1_data[i] = value if value is not None else 0

    for i, value in enumerate(m_data[:end_qty]):
        massive2_data[i] = value if value is not None else 0

    for i, value in enumerate(string_data):
        string2_data[i] = value if value is not None else ""

    write_results_massive1 = []
    write_results_massive2 = []
    write_results_massive5 = []
    write_results_massive4 = []
    write_results_string2 = []

    with LogixDriver(PLC_ADDRESS) as plc:
        for i in range(1, 121):
            length_tag = f'massive1[{i}]'
            qty_tag = f'massive2[{i}]'
            massive5_tag = f'massive5[{i}]'
            massive4_tag = f'massive4[{i}]'

            write_length = plc.write(length_tag, massive1_data[i - 1])
            write_qty = plc.write(qty_tag, massive2_data[i - 1])
            write_massive5 = plc.write(massive5_tag, 0)
            write_massive4 = plc.write(massive4_tag, 0)

            write_results_massive1.append(write_length)
            write_results_massive2.append(write_qty)
            write_results_massive5.append(write_massive5)
            write_results_massive4.append(write_massive4)

        for i in range(1, 121):
            tag_name = f'string2[{i}]'
            string_value = string2_data[i - 1]

            write_string = plc.write(tag_name, string_value)
            write_results_string2.append(write_string)

    result_str = f'Write successful for massive1: {", ".join(str(val.value) for val in write_results_massive1)}\n'
    result_str += f'Write successful for massive2: {", ".join(str(val.value) for val in write_results_massive2)}\n'
    result_str += f'Write successful for massive5: {", ".join(str(val.value) for val in write_results_massive5)}\n'
    result_str += f'Write successful for massive4: {", ".join(str(val.value) for val in write_results_massive4)}\n'
    result_str += f'Write successful for string2: {", ".join(str(val.value) for val in write_results_string2)}'

    messagebox.showinfo("Write Results", result_str)


def show_about():
    about_text = """PLC Data Writer V3

This program is designed for specific use with a particular machine and a specific Excel table. It should only be used by authorized personnel, as it can lead to changes in the machine's behavior, potential damage to the machine, and even harm to personnel.

Special Thanks to Library Creators:
- This program makes use of various libraries, and we express our gratitude to their creators for their contributions.

Author:
- Stanimir Georgiev
- Email: stereophonek@gmail.com

Distribution:
- This program should only be distributed by Stanimir Georgiev.
"""
    messagebox.showinfo("About", about_text)


def show_help():
    help_text = """PLC Data Writer Help

How the Program Works:
1. Enter the path to your Excel file containing data.
2. Set the 'Length End,' 'Qty End,' and 'Order No End' values:
   - 'Length End' specifies the endpoint for the length data in column 'J'.
   - 'Qty End' specifies the endpoint for the quantity data in column 'I'.
   - 'Order No End' specifies the endpoint for the order numbers in column 'B'.
   Please note that these values are specific to your Excel table structure.

3. Check the 'Check if the machine is in auto mode' checkbox if needed.
4. Click the 'Write to PLC' button to write the data to the PLC.

Endpoint Values:
- 'Length End' starts at column 'J2' and allows up to 120 rows (J2 to J61).
- 'Qty End' starts at column 'I2' and allows up to 120 rows (I2 to I61).
- 'Order No End' starts at 'B2' and allows up to 120 rows (B2 to B61).

Please make sure to set these values correctly based on your specific Excel data structure.

Use this program responsibly and only with authorization to avoid any undesirable consequences.
"""
    messagebox.showinfo("Help", help_text)


root = tk.Tk()
root.title("PLC Data Writer V2")
root.geometry("400x400")
set_icon(your_icon_base64_code)

plc_address_label = tk.Label(root, text=f"PLC Address: {PLC_ADDRESS}")
plc_address_label.pack()

file_path_label = tk.Label(root, text="Excel File:")
file_path_label.pack()

file_path_entry = tk.Entry(root)
file_path_entry.pack()

browse_button = tk.Button(root, text="Browse", command=browse_file)
browse_button.pack()

length_end_label = tk.Label(root, text="Length End (J column,start at 2, max 121):")
length_end_label.pack()
length_end_entry = tk.Entry(root)
length_end_entry.pack()

qty_end_label = tk.Label(root, text="Qty End (I column, start at 2, max 121):")
qty_end_label.pack()
qty_end_entry = tk.Entry(root)
qty_end_entry.pack()

string_end_label = tk.Label(root, text="Order No End (B column,B2=2, max 121):")
string_end_label.pack()
string_end_entry = tk.Entry(root)
string_end_entry.pack()

automated_mode = tk.IntVar(value=1)  # Initially checked
automated_mode_checkbox = tk.Checkbutton(root, text="Check if the machine is in auto mode", variable=automated_mode)
automated_mode_checkbox.pack()

write_button = tk.Button(root, text="Write to PLC", command=check_and_write_to_plc)
write_button.pack()

menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

about_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="About", menu=about_menu)
about_menu.add_command(label="About", command=show_about)
help_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Help", menu=help_menu)
help_menu.add_command(label="Help", command=show_help)
root.mainloop()
